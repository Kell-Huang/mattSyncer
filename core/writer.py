import os
import re
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import polars as pl
import xlsxwriter
from openpyxl import load_workbook
from PySide6.QtCore import QObject, Signal

from utils.encoding import detect_encoding


class Writer(QObject):
    """
    Writes updated data to a new file with minimal memory usage.

    Supports XLSX (via xlsxwriter for speed) and CSV output.
    Preserves header formatting and column widths (XLSX only).
    Formula columns can be explicitly specified via session_data to improve
    performance and guarantee formula preservation; otherwise they are
    auto‑detected from the first 100 rows of the target file.
    New rows (additions) receive the formula from the last row of the
    target file with relative references adjusted (fill‑down behaviour).
    Creates a timestamped backup of the original output file before
    overwriting it.
    All non‑SKU columns are converted to Utf8 to prevent type conflicts
    during updates and additions.
    """

    progress_updated = Signal(str)

    FORMULA_SCAN_ROWS = 100
    TARGET_FORMULA_SCAN_ROWS = 100

    def __init__(
        self,
        session_data: Dict[str, Any],
        diff_result: Dict[str, Any],
        result_df: Optional[pl.DataFrame] = None,
    ):
        """Initialize the writer with session and diff data."""
        super().__init__()
        self.source_path = session_data.get("source_path", "")
        self.target_path = session_data["target_path"]
        self.output_path = session_data.get("output_path", self.target_path)
        self.column_mapping = session_data["column_mapping"]
        self.target_sku_col = session_data["target_sku_col"]

        self._user_formula_cols: Optional[List[Union[str, int]]] = (
            session_data.get("formula_columns")
        )
        self.formula_info = session_data.get("formula_info", {})

        self._formula_matrix: Optional[List[List[Optional[str]]]] = None
        self._formula_col_indices_from_target: Set[int] = set()
        self._target_data_rows: int = 0

        self.updates = diff_result.get("updates", [])
        self.additions = diff_result.get("additions")
        self.new_columns_data = diff_result.get("new_columns_data", {})

        self._result_df = result_df

        self._build_update_lookup()
        self._build_new_columns_lookup()
        self.formula_cells_written: Set[Tuple[int, int]] = set()

    def _build_update_lookup(self):
        """Build lookup dictionary for efficient row updates."""
        self.updates_by_sku: Dict[str, Dict[str, Any]] = {}
        self.sku_updated_columns: Dict[str, Set[str]] = {}
        for update in self.updates:
            sku = update["sku"]
            col = update["column"]
            if sku not in self.updates_by_sku:
                self.updates_by_sku[sku] = {}
                self.sku_updated_columns[sku] = set()
            self.updates_by_sku[sku][col] = update["new_value"]
            self.sku_updated_columns[sku].add(col)

        self.update_skus = set(self.updates_by_sku.keys())
        self.update_columns = list(
            self.column_mapping.get("updates", {}).values()
        )

    def _build_new_columns_lookup(self):
        """Build lookup dictionary for new column values by row index."""
        new_columns_mapping = self.column_mapping.get("new_columns", {})
        self.new_columns_combined = None
        if not new_columns_mapping or not self.new_columns_data:
            return

        dfs = list(self.new_columns_data.values())
        combined = dfs[0]
        for df in dfs[1:]:
            combined = combined.join(df, on="row_idx", how="left")
        self.new_columns_combined = combined

    def write(self):
        """Execute the complete writing process."""
        is_csv = self.output_path.endswith(".csv")

        if is_csv:
            temp_path = self.output_path.replace(".csv", "_temp.csv")
        else:
            temp_path = self.output_path.replace(".xlsx", "_temp.xlsx")

        self.formula_cells_written = set()
        self._formula_matrix = None
        self._formula_col_indices_from_target = set()
        self._target_data_rows = 0

        try:
            if is_csv:
                self._write_csv(temp_path)
            else:
                self._write_xlsx(temp_path)

            if os.path.exists(self.output_path):
                self._create_backup(self.output_path)
                os.remove(self.output_path)

            shutil.move(temp_path, self.output_path)

        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e

        finally:
            self.updates_by_sku.clear()
            self._formula_matrix = None

    def _create_backup(self, file_path: str):
        """Create a timestamped backup of the given file in the same directory."""
        base, ext = os.path.splitext(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{base}_backup_{timestamp}{ext}"
        shutil.copy2(file_path, backup_path)
        self.progress_updated.emit(f"Backup created: {backup_path}")

    # ── Formula extraction helpers ──────────────────────────────────────
    def _resolve_target_formula_cols(self, headers: List[str]) -> Set[int]:
        """Determine formula column indices (0‑based)."""
        if self._user_formula_cols is not None:
            formula_cols = set()
            for item in self._user_formula_cols:
                if isinstance(item, int):
                    if 0 <= item < len(headers):
                        formula_cols.add(item)
                elif isinstance(item, str):
                    if item in headers:
                        formula_cols.add(headers.index(item))
            return formula_cols

        if not self.target_path.lower().endswith(".xlsx"):
            return set()

        formula_cols = set()
        wb = load_workbook(self.target_path, read_only=True, data_only=False)
        ws = wb.active
        try:
            for row_idx, row in enumerate(ws.iter_rows()):
                if row_idx == 0:
                    continue
                if row_idx > self.TARGET_FORMULA_SCAN_ROWS:
                    break
                for col_idx, cell in enumerate(row):
                    try:
                        if cell.data_type == "f":
                            formula_cols.add(col_idx)
                    except AttributeError:
                        pass
        finally:
            wb.close()
        return formula_cols

    def _build_formula_matrix(self, formula_cols: Set[int]):
        """Build formula matrix from target file."""
        if not formula_cols:
            self._formula_matrix = None
            self._target_data_rows = 0
            return

        sorted_cols = sorted(formula_cols)
        col_to_idx = {col: idx for idx, col in enumerate(sorted_cols)}
        num_formula_cols = len(sorted_cols)

        matrix: List[List[Optional[str]]] = []

        wb = load_workbook(self.target_path, read_only=True, data_only=False)
        ws = wb.active
        try:
            for row_idx, row in enumerate(ws.iter_rows()):
                if row_idx == 0:
                    continue
                row_formulas: List[Optional[str]] = [None] * num_formula_cols
                for col_idx, cell in enumerate(row):
                    if col_idx in formula_cols:
                        try:
                            if cell.data_type == "f":
                                row_formulas[col_to_idx[col_idx]] = cell.value
                        except AttributeError:
                            pass
                matrix.append(row_formulas)
        finally:
            wb.close()

        self._formula_matrix = matrix
        self._target_data_rows = len(matrix)
        self._formula_col_indices_from_target = formula_cols

    # ── Apply updates via pivot ────────────────────────────────────────
    def _apply_updates(self, source_lf):
        """Apply all column updates in a single join using Polars pivot.
        All update values are cast to Utf8 to avoid type conflicts."""
        if not self.updates:
            return source_lf

        # Normalise update values: convert non‑None values to strings
        for u in self.updates:
            if u.get("new_value") is not None:
                u["new_value"] = str(u["new_value"])
            if u.get("old_value") is not None:
                u["old_value"] = str(u["old_value"])

        sku_col = self.target_sku_col
        updates_df = pl.DataFrame(self.updates)
        # Ensure update values and SKU are strings
        updates_df = updates_df.with_columns(
            [pl.col("sku").cast(pl.Utf8), pl.col("new_value").cast(pl.Utf8)]
        )

        updates_wide = (
            updates_df.pivot(
                index="sku",
                columns="column",
                values="new_value",
                aggregate_function="last",
            )
            .rename({"sku": sku_col})
            .lazy()
        )

        source_lf = source_lf.with_columns(pl.col(sku_col).cast(pl.Utf8))

        update_cols = [
            c for c in updates_wide.collect_schema().names() if c != sku_col
        ]
        rename_map = {c: f"__upd_{c}" for c in update_cols}
        updates_wide = updates_wide.rename(rename_map)
        source_lf = source_lf.join(updates_wide, on=sku_col, how="left")

        coalesce_exprs = []
        source_schema = source_lf.collect_schema()
        for col in update_cols:
            upd_col = f"__upd_{col}"
            if upd_col in source_schema.names():
                coalesce_exprs.append(
                    pl.when(pl.col(upd_col).is_not_null())
                    .then(pl.col(upd_col))
                    .otherwise(pl.col(col))
                    .alias(col)
                )
        if coalesce_exprs:
            source_lf = source_lf.with_columns(coalesce_exprs)
        temp_cols = [
            c
            for c in source_lf.collect_schema().names()
            if c.startswith("__upd_")
        ]
        if temp_cols:
            source_lf = source_lf.drop(temp_cols)
        return source_lf

    # ── XLSX Writing ────────────────────────────────────────────────────
    def _get_header_formats(self):
        """Extract header formatting from the original target file."""
        if not self.target_path.lower().endswith(
            (".xlsx", ".xlsm", ".xltx", ".xltm")
        ):
            return {}

        wb = load_workbook(self.target_path, read_only=True)
        ws = wb.active
        header_formats = {}

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)

            font = cell.font
            fill = cell.fill
            align = cell.alignment

            font_color = None
            try:
                if font.color and font.color.rgb:
                    raw = str(font.color.rgb)
                    if raw and raw != "00000000":
                        if len(raw) == 8:
                            raw = raw[2:]
                        if len(raw) == 6:
                            font_color = f"#{raw}"
            except (AttributeError, ValueError, TypeError):
                pass

            bg_color = None
            try:
                if fill and fill.start_color and fill.start_color.rgb:
                    raw = str(fill.start_color.rgb)
                    if raw and raw != "00000000":
                        if len(raw) == 8:
                            raw = raw[2:]
                        if len(raw) == 6:
                            bg_color = f"#{raw}"
            except (AttributeError, ValueError, TypeError):
                pass

            num_format = None
            try:
                num_format = cell.number_format
            except AttributeError:
                pass

            header_formats[col_idx] = {
                "font_name": font.name if font.name else "Calibri",
                "font_size": font.size if font.size else 11,
                "bold": bool(font.bold),
                "font_color": font_color,
                "bg_color": bg_color,
                "align": (
                    align.horizontal if align and align.horizontal else "left"
                ),
                "valign": (
                    align.vertical if align and align.vertical else "bottom"
                ),
                "num_format": num_format,
            }

        wb.close()
        return header_formats

    def _write_xlsx(self, output_path: str):
        """Write data to XLSX using write_column for all columns, then overlay formulas."""
        self.progress_updated.emit("Building output data...")
        source_lf = self._get_lazy_source()
        source_lf = self._apply_updates(source_lf)

        # Add new columns
        new_columns_mapping = self.column_mapping.get("new_columns", {})
        if new_columns_mapping and self.new_columns_combined is not None:
            new_cols_lf = self.new_columns_combined.lazy()
            source_lf = source_lf.with_row_index(name="_row_idx_")
            source_lf = source_lf.join(
                new_cols_lf.rename({"row_idx": "_row_idx_"}),
                on="_row_idx_",
                how="left",
            )
            source_lf = source_lf.drop("_row_idx_")

        # Append additions
        if self.additions is not None and len(self.additions) > 0:
            add_lf = self.additions.lazy()
            source_schema = source_lf.collect_schema()
            source_cols = source_schema.names()
            add_schema = add_lf.collect_schema()
            for col in source_cols:
                if col not in add_schema.names():
                    add_lf = add_lf.with_columns(pl.lit(None).alias(col))
            add_lf = add_lf.select(source_cols)
            for col in source_cols:
                if col in add_lf.collect_schema().names():
                    if (
                        add_lf.collect_schema()[col] != source_schema[col]
                        and add_lf.collect_schema()[col] != pl.Null
                    ):
                        add_lf = add_lf.with_columns(
                            pl.col(col).cast(source_schema[col], strict=False)
                        )
            source_lf = pl.concat([source_lf, add_lf])

        # Direct collect – eliminates temporary Parquet I/O
        result_df = source_lf.collect()
        del source_lf

        total_rows = len(result_df)
        headers = result_df.columns
        self.progress_updated.emit(f"Writing {total_rows:,} rows to XLSX...")

        # Resolve formula columns
        target_formula_cols = self._resolve_target_formula_cols(headers)
        sample_rows = min(self.FORMULA_SCAN_ROWS, total_rows)
        sample = result_df.head(sample_rows)
        data_formula_cols = self._detect_formula_columns(sample)
        all_formula_cols = target_formula_cols | data_formula_cols

        user_formula_cols_indices = set()
        if self._user_formula_cols is not None:
            for item in self._user_formula_cols:
                if isinstance(item, int) and 0 <= item < len(headers):
                    user_formula_cols_indices.add(item)
                elif isinstance(item, str) and item in headers:
                    user_formula_cols_indices.add(headers.index(item))

        if target_formula_cols:
            self._build_formula_matrix(target_formula_cols)
        else:
            self._formula_matrix = None
            self._target_data_rows = 0

        self.progress_updated.emit(
            f"Formula columns detected: {len(all_formula_cols)} "
            f"(target: {len(target_formula_cols)}, data: {len(data_formula_cols)})"
        )

        header_formats = self._get_header_formats()
        col_index_by_name = {name: idx for idx, name in enumerate(headers)}
        sku_col_idx = col_index_by_name.get(self.target_sku_col)

        sku_to_updated_col_indices: Dict[Any, Set[int]] = {}
        for sku, updated_cols in self.sku_updated_columns.items():
            indices = set()
            for col_name in updated_cols:
                if col_name in col_index_by_name:
                    indices.add(col_index_by_name[col_name])
            if indices:
                sku_to_updated_col_indices[sku] = indices

        formula_matrix = self._formula_matrix
        formula_target_cols = self._formula_col_indices_from_target
        target_data_rows = self._target_data_rows

        if formula_matrix is not None and target_data_rows > 0:
            sorted_target_cols = sorted(formula_target_cols)
            col_to_matrix_idx = {
                col: idx for idx, col in enumerate(sorted_target_cols)
            }
            last_row_formulas = formula_matrix[-1]
        else:
            col_to_matrix_idx = {}
            last_row_formulas = None

        _new_col_target_indices = None
        if new_columns_mapping and user_formula_cols_indices:
            _new_col_target_indices = []
            for new_col_name in new_columns_mapping.values():
                if new_col_name in headers:
                    col_idx = headers.index(new_col_name)
                    col_letter = self._column_index_to_letter(col_idx)
                    _new_col_target_indices.append((col_idx, col_letter))

        workbook = xlsxwriter.Workbook(output_path, {"constant_memory": False})
        worksheet = workbook.add_worksheet()

        fmt_cache = {}

        def get_fmt(col_idx):
            if col_idx not in header_formats:
                return None
            props = header_formats[col_idx]
            key = tuple(sorted(props.items()))
            if key not in fmt_cache:
                fmt_kwargs = {
                    "font_name": props["font_name"],
                    "font_size": props["font_size"],
                    "bold": props["bold"],
                    "align": props["align"],
                    "valign": props["valign"],
                    "border": 1,
                }
                if props.get("font_color"):
                    fmt_kwargs["font_color"] = props["font_color"]
                if props.get("bg_color"):
                    fmt_kwargs["bg_color"] = props["bg_color"]
                if props.get("num_format"):
                    fmt_kwargs["num_format"] = props["num_format"]
                fmt = workbook.add_format(fmt_kwargs)
                fmt_cache[key] = fmt
            return fmt_cache[key]

        for col_idx, header in enumerate(headers):
            fmt = get_fmt(col_idx + 1)
            if fmt:
                worksheet.write(0, col_idx, header, fmt)
            else:
                worksheet.write(0, col_idx, header)

        self.progress_updated.emit(f"Writing {len(headers)} columns...")
        for col_idx, col_name in enumerate(headers):
            col_series = result_df[col_name]
            col_data = col_series.to_list()
            worksheet.write_column(1, col_idx, col_data)
            del col_series, col_data

        if all_formula_cols:
            self.progress_updated.emit(
                f"Overlaying formulas on {len(all_formula_cols)} columns..."
            )
            _worksheet = worksheet
            _user_formula_cols_indices = user_formula_cols_indices
            _col_to_matrix_idx = col_to_matrix_idx
            _formula_matrix = formula_matrix
            _target_data_rows = target_data_rows
            _last_row_formulas = last_row_formulas
            _sku_col_idx = sku_col_idx
            _sku_to_updated_col_indices = sku_to_updated_col_indices

            chunk_size = 20000
            for start in range(0, total_rows, chunk_size):
                end = min(start + chunk_size, total_rows)
                batch = result_df[start:end]

                for rel_idx, row_data in enumerate(batch.iter_rows()):
                    row_idx = start + rel_idx
                    r = row_idx + 1

                    for col_idx in all_formula_cols:
                        val = row_data[col_idx]
                        write_formula = False
                        formula_text = None
                        restored_from_matrix = False

                        if isinstance(val, str) and val.startswith("="):
                            write_formula = True
                            formula_text = val
                        elif (
                            col_idx in _col_to_matrix_idx
                            and _formula_matrix is not None
                        ):
                            is_formula_update = False
                            if (
                                _sku_col_idx is not None
                                and _sku_to_updated_col_indices
                            ):
                                sku_value = row_data[_sku_col_idx]
                                updated_cols = _sku_to_updated_col_indices.get(
                                    sku_value
                                )
                                if updated_cols and col_idx in updated_cols:
                                    new_val = val
                                    if isinstance(
                                        new_val, str
                                    ) and new_val.startswith("="):
                                        is_formula_update = True
                            if not is_formula_update:
                                if row_idx < _target_data_rows:
                                    matrix_idx = _col_to_matrix_idx[col_idx]
                                    orig_formula = _formula_matrix[row_idx][
                                        matrix_idx
                                    ]
                                    if orig_formula is not None:
                                        write_formula = True
                                        formula_text = orig_formula
                                        restored_from_matrix = True
                                else:
                                    if _last_row_formulas is not None:
                                        matrix_idx = _col_to_matrix_idx[
                                            col_idx
                                        ]
                                        fill_formula = _last_row_formulas[
                                            matrix_idx
                                        ]
                                        if fill_formula is not None:
                                            row_offset = (row_idx + 2) - (
                                                _target_data_rows + 1
                                            )
                                            if row_offset != 0:
                                                fill_formula = (
                                                    self._offset_formula_row(
                                                        fill_formula,
                                                        row_offset,
                                                    )
                                                )
                                            write_formula = True
                                            formula_text = fill_formula
                                            restored_from_matrix = True

                        if (
                            restored_from_matrix
                            and col_idx in _user_formula_cols_indices
                            and _new_col_target_indices is not None
                        ):
                            suffix_parts = []
                            for _, col_letter in _new_col_target_indices:
                                suffix_parts.append(
                                    f'&";"&${col_letter}$1&":"&{col_letter}{r}'
                                )
                            new_suffix = "".join(suffix_parts)
                            formula_text = formula_text + new_suffix

                        if write_formula:
                            _worksheet.write_formula(r, col_idx, formula_text)

                if start % (chunk_size * 10) == 0:
                    self.progress_updated.emit(
                        f"Written {min(end, total_rows):,}/{total_rows:,} rows"
                    )

        workbook.close()
        del result_df

        self.progress_updated.emit("XLSX write complete.")

    def _detect_formula_columns(self, sample: pl.DataFrame) -> Set[int]:
        """Identify columns containing text formulas."""
        formula_cols = set()
        for col_idx, col_name in enumerate(sample.columns):
            series = sample[col_name]
            if series.dtype == pl.Utf8 and series.str.starts_with("=").any():
                formula_cols.add(col_idx)
        return formula_cols

    @staticmethod
    def _offset_formula_row(formula: str, row_offset: int) -> str:
        """Adjust relative row references in a formula."""
        pattern = re.compile(r"(\$?[A-Za-z]+)(\$?)(\d+)")

        def repl(m):
            col_part = m.group(1)
            row_abs = m.group(2)
            row_num = int(m.group(3))
            if row_abs == "":
                return f"{col_part}{row_num + row_offset}"
            else:
                return m.group(0)

        return pattern.sub(repl, formula)

    # ── CSV Writing ─────────────────────────────────────────────────────
    def _write_csv(self, output_path: str):
        """Write data to CSV with UTF‑8 BOM."""
        self.progress_updated.emit("Building output data...")
        source_lf = self._get_lazy_source()
        source_lf = self._apply_updates(source_lf)

        new_columns_mapping = self.column_mapping.get("new_columns", {})
        if new_columns_mapping and self.new_columns_combined is not None:
            new_cols_lf = self.new_columns_combined.lazy()
            source_lf = source_lf.with_row_index(name="_row_idx_")
            source_lf = source_lf.join(
                new_cols_lf.rename({"row_idx": "_row_idx_"}),
                on="_row_idx_",
                how="left",
            )
            source_lf = source_lf.drop("_row_idx_")

        if self.additions is not None and len(self.additions) > 0:
            add_lf = self.additions.lazy()
            source_schema = source_lf.collect_schema()
            source_cols = source_schema.names()
            add_schema = add_lf.collect_schema()
            for col in source_cols:
                if col not in add_schema.names():
                    add_lf = add_lf.with_columns(pl.lit(None).alias(col))
            add_lf = add_lf.select(source_cols)
            for col in source_cols:
                if col in add_lf.collect_schema().names():
                    if (
                        add_lf.collect_schema()[col] != source_schema[col]
                        and add_lf.collect_schema()[col] != pl.Null
                    ):
                        add_lf = add_lf.with_columns(
                            pl.col(col).cast(source_schema[col], strict=False)
                        )
            source_lf = pl.concat([source_lf, add_lf])

        self.progress_updated.emit("Collecting result...")
        result_df = source_lf.collect()
        self.progress_updated.emit(
            f"Writing {len(result_df):,} rows to CSV..."
        )

        temp_csv = output_path + ".tmp"
        try:
            result_df.write_csv(temp_csv, batch_size=20000)
            del result_df
            self.progress_updated.emit("Adding UTF‑8 BOM...")
            with open(temp_csv, "rb") as src, open(output_path, "wb") as dst:
                dst.write(b"\xef\xbb\xbf")
                shutil.copyfileobj(src, dst)
        finally:
            try:
                os.remove(temp_csv)
            except OSError:
                pass

        self.progress_updated.emit("CSV write complete.")

    def _get_lazy_source(self):
        """Obtain a LazyFrame from the target file, or use the pre-built result_df."""
        if self._result_df is not None:
            source_lf = self._result_df.lazy()
        elif self.target_path.endswith(".csv"):

            encoding = detect_encoding(self.target_path)

            # Read header only to get raw column names
            header_df = pl.read_csv(
                self.target_path,
                has_header=True,
                encoding=encoding,
                n_rows=0,
                infer_schema_length=0,
            )
            raw_columns = header_df.columns
            clean_columns = [
                str(c).lstrip("\ufeff").strip() for c in raw_columns
            ]
            raw_to_clean = dict(zip(raw_columns, clean_columns))

            # Force all columns to Utf8
            schema_overrides = {raw_col: pl.Utf8 for raw_col in raw_columns}

            source_lf = pl.scan_csv(
                self.target_path,
                has_header=True,
                encoding=encoding,
                truncate_ragged_lines=True,
                schema_overrides=schema_overrides,
            ).rename(raw_to_clean)
        else:
            df = pl.read_excel(self.target_path, engine="calamine")
            df.columns = [str(c).strip() for c in df.columns]
            source_lf = df.lazy()

        # For Excel sources, cast non-SKU columns to Utf8
        if not self.target_path.endswith(".csv"):
            non_sku_cols = [
                c
                for c in source_lf.collect_schema().names()
                if c != self.target_sku_col
            ]
            if non_sku_cols:
                source_lf = source_lf.with_columns(
                    [
                        pl.col(c).cast(pl.Utf8, strict=False)
                        for c in non_sku_cols
                    ]
                )

        return source_lf

    @staticmethod
    def _column_index_to_letter(col_idx: int) -> str:
        """Convert a 0‑based column index to Excel column letters."""
        letters = []
        while col_idx >= 0:
            remainder = col_idx % 26
            letters.append(chr(ord("A") + remainder))
            col_idx = col_idx // 26 - 1
        return "".join(reversed(letters))
