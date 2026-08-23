from typing import Any, Dict, List, Optional

import polars as pl

from utils.encoding import detect_encoding


class Differ:
    """
    Compares source and target data to find differences.

    Uses Polars for efficient join operations and column-wise comparison.
    Handles both existing column updates and new column addition.
    """

    def __init__(
        self,
        session_data: Dict[str, Any],
        target_df: Optional[pl.DataFrame] = None,
        sheet_name: Optional[str] = None,
    ):
        """Initialize the differ with session configuration.

        Args:
            session_data: Dictionary containing:
                - source_path: Source file path
                - target_path: Target file path (ignored if target_df is provided)
                - source_sku_col, target_sku_col: SKU column names
                - column_mapping: Dict with 'updates' and 'new_columns'
                - output_path: Output file path
            target_df: Optional pre-loaded target DataFrame. If provided,
                      _read_target will be skipped and this DataFrame used instead.
            sheet_name: Optional sheet name to read from an Excel file.
                       If None, the first sheet is read (default behavior).
        """
        self.source_path = session_data["source_path"]
        self.target_path = session_data.get("target_path", "")
        self.source_sku_col = session_data["source_sku_col"]
        self.target_sku_col = session_data["target_sku_col"]
        self.column_mapping = session_data["column_mapping"]
        self.output_path = session_data.get("output_path", "")

        self.target_df = target_df  # None means read from disk
        self.sheet_name = (
            sheet_name  # Specific sheet to read (None = first sheet)
        )

        self.source_df = None
        self.updates = []
        self.additions = None
        self.new_columns_data = {}
        self.warnings = []

    def compare(self) -> Dict[str, Any]:
        """Perform the full comparison between source and target.

        Returns:
            Dictionary with comparison results.
            If the source sheet is missing the SKU column, an empty result
            with a warning is returned.
        """
        # Read source file
        self._read_source()

        # If the source sheet does not contain the required SKU column,
        # return an empty result with a warning (the sheet will be skipped)
        if self.source_df is None:
            return {
                "updates": [],
                "additions": None,
                "new_columns_data": {},
                "warnings": self.warnings,
                "source_rows": 0,
                "target_rows": (
                    len(self.target_df) if self.target_df is not None else 0
                ),
                "columns_matched": len(self.column_mapping.get("updates", {})),
                "new_columns_count": len(
                    self.column_mapping.get("new_columns", {})
                ),
                "new_skus": 0,
            }

        # Read target file only if not already provided
        if self.target_df is None:
            self._read_target()
        # else: target_df already set, skip reading

        # Clean and deduplicate source data
        self._clean_source()
        self._deduplicate_source()

        # Perform join and comparison operations
        self._perform_comparison()

        return {
            "updates": self.updates,
            "additions": self.additions,
            "new_columns_data": self.new_columns_data,
            "warnings": self.warnings,
            "source_rows": len(self.source_df),
            "target_rows": (
                len(self.target_df) if self.target_df is not None else 0
            ),
            "columns_matched": len(self.column_mapping.get("updates", {})),
            "new_columns_count": len(
                self.column_mapping.get("new_columns", {})
            ),
            "new_skus": (
                len(self.additions) if self.additions is not None else 0
            ),
        }

    def _read_source(self):
        """Read source file into DataFrame, only loading required columns."""
        update_src_cols = list(self.column_mapping.get("updates", {}).keys())
        new_src_cols = list(self.column_mapping.get("new_columns", {}).keys())
        required_src_cols = list(
            dict.fromkeys(
                update_src_cols + new_src_cols + [self.source_sku_col]
            )
        )

        if self.source_path.endswith(".csv"):
            from utils.encoding import detect_encoding

            encoding = detect_encoding(self.source_path)

            # Read header only to get raw column names
            header_df = pl.read_csv(
                self.source_path,
                has_header=True,
                encoding=encoding,
                n_rows=0,
                infer_schema_length=0,
            )
            raw_columns = header_df.columns
            clean_columns = [
                str(c).lstrip("\ufeff").strip() for c in raw_columns
            ]
            col_map = {
                clean: raw for clean, raw in zip(clean_columns, raw_columns)
            }

            existing_clean = [c for c in required_src_cols if c in col_map]
            if self.source_sku_col not in existing_clean:
                self.warnings.append(
                    {
                        "type": "missing_sku",
                        "message": f"Sheet '{self.sheet_name or 'first sheet'}' in {self.source_path} "
                        f"does not contain SKU column '{self.source_sku_col}'. Skipped.",
                    }
                )
                self.source_df = None
                return

            raw_existing = [col_map[c] for c in existing_clean]
            schema_overrides = {raw_col: pl.Utf8 for raw_col in raw_existing}

            self.source_df = pl.read_csv(
                self.source_path,
                has_header=True,
                encoding=encoding,
                columns=raw_existing,
                truncate_ragged_lines=True,
                schema_overrides=schema_overrides,
            )

            # Rename columns based on actual raw names returned by polars
            raw_to_clean = {
                raw: clean for raw, clean in zip(raw_existing, existing_clean)
            }
            current_cols = self.source_df.columns
            self.source_df.columns = [
                raw_to_clean.get(c, c) for c in current_cols
            ]
        else:
            if self.sheet_name is not None:
                self.source_df = pl.read_excel(
                    self.source_path,
                    sheet_name=self.sheet_name,
                    engine="calamine",
                )
            else:
                self.source_df = pl.read_excel(
                    self.source_path, engine="calamine"
                )
            self.source_df.columns = [
                str(c).lstrip("\ufeff").strip() for c in self.source_df.columns
            ]

        if (
            self.source_df is not None
            and self.source_sku_col not in self.source_df.columns
        ):
            self.warnings.append(
                {
                    "type": "missing_sku",
                    "message": f"Sheet '{self.sheet_name or 'first sheet'}' in {self.source_path} "
                    f"does not contain SKU column '{self.source_sku_col}'. Skipped.",
                }
            )
            self.source_df = None

    def _read_target(self):
        """Read target file efficiently - only needed columns."""
        update_columns = list(self.column_mapping.get("updates", {}).values())
        if self.target_sku_col not in update_columns:
            update_columns.append(self.target_sku_col)
        required_target_cols = list(dict.fromkeys(update_columns))

        if self.target_path.endswith(".csv"):
            from utils.encoding import detect_encoding

            encoding = detect_encoding(self.target_path)

            # Read header only to get raw column names
            header_df = pl.read_csv(
                self.target_path,
                has_header=True,
                encoding=encoding,
                n_rows=0,
                infer_schema_length=0,
            )
            raw_columns = header_df.columns
            clean_columns = [
                str(c).lstrip("\ufeff").strip() for c in raw_columns
            ]
            col_map = {
                clean: raw for clean, raw in zip(clean_columns, raw_columns)
            }

            existing_clean = [c for c in required_target_cols if c in col_map]
            if self.target_sku_col not in existing_clean:
                raise ValueError(
                    f"Target SKU column '{self.target_sku_col}' not found in target file"
                )

            raw_existing = [col_map[c] for c in existing_clean]
            schema_overrides = {raw_col: pl.Utf8 for raw_col in raw_existing}

            self.target_df = pl.read_csv(
                self.target_path,
                has_header=True,
                encoding=encoding,
                columns=raw_existing,
                truncate_ragged_lines=True,
                schema_overrides=schema_overrides,
            )

            # Rename columns based on actual raw names returned by polars
            raw_to_clean = {
                raw: clean for raw, clean in zip(raw_existing, existing_clean)
            }
            current_cols = self.target_df.columns
            self.target_df.columns = [
                raw_to_clean.get(c, c) for c in current_cols
            ]
        else:
            from openpyxl import load_workbook

            # Read only the header row to get raw column names
            wb = load_workbook(self.target_path, read_only=True)
            ws = wb.active
            header_row = next(
                ws.iter_rows(min_row=1, max_row=1, values_only=True), None
            )
            wb.close()
            if header_row is None:
                raise ValueError(
                    f"Target file {self.target_path} has no header row"
                )

            raw_columns = [str(c).strip() for c in header_row]
            clean_columns = [
                str(c).lstrip("\ufeff").strip() for c in raw_columns
            ]
            col_map = {
                clean: raw for clean, raw in zip(clean_columns, raw_columns)
            }

            existing_clean = [c for c in required_target_cols if c in col_map]
            if self.target_sku_col not in existing_clean:
                raise ValueError(
                    f"Target SKU column '{self.target_sku_col}' not found in target file"
                )

            raw_existing = [col_map[c] for c in existing_clean]

            self.target_df = pl.read_excel(
                self.target_path,
                engine="calamine",
                columns=raw_existing,
            )

            # Rename columns based on actual raw names returned by polars
            raw_to_clean = {
                raw: clean for raw, clean in zip(raw_existing, existing_clean)
            }
            current_cols = self.target_df.columns
            self.target_df.columns = [
                raw_to_clean.get(c, c) for c in current_cols
            ]

    def _clean_source(self):
        """Clean source data - remove checkmarks and empty values."""
        all_source_cols = list(
            self.column_mapping.get("updates", {}).keys()
        ) + list(self.column_mapping.get("new_columns", {}).keys())

        cleaning_exprs = []
        for source_col in all_source_cols:
            if source_col in self.source_df.columns:
                cleaning_exprs.append(
                    pl.when(
                        pl.col(source_col)
                        .cast(pl.Utf8)
                        .str.strip_chars()
                        .is_in(
                            ["✔", "✓", "√", "✅", "☑", "", " ", "N/A", "n/a"]
                        )
                    )
                    .then(None)
                    .otherwise(pl.col(source_col))
                    .alias(source_col)
                )

        if cleaning_exprs:
            self.source_df = self.source_df.with_columns(cleaning_exprs)

    def _deduplicate_source(self):
        """Deduplicate source data by SKU, keeping first occurrence."""
        sku_col = self.source_sku_col

        # Find duplicates
        duplicates = (
            self.source_df.group_by(sku_col).len().filter(pl.col("len") > 1)
        )

        if len(duplicates) > 0:
            dup_skus = duplicates[sku_col].to_list()
            self.warnings.append(
                {
                    "type": "duplicate_sku",
                    "message": (
                        f"Found {len(dup_skus)} duplicate SKUs in source file. "
                        f"Keeping first occurrence."
                    ),
                    "skus": dup_skus,
                    "count": len(dup_skus),
                }
            )

            # Keep only first occurrence of each SKU
            self.source_df = self.source_df.unique(
                subset=[sku_col], keep="first"
            )

    def warn_target_duplicates(self):
        """Add warning if target file contains duplicate SKU values."""
        if (
            self.target_df is None
            or self.target_sku_col not in self.target_df.columns
        ):
            return

        duplicates = (
            self.target_df.group_by(self.target_sku_col)
            .len()
            .filter(pl.col("len") > 1)
        )
        if len(duplicates) > 0:
            dup_skus = duplicates[self.target_sku_col].to_list()
            self.warnings.append(
                {
                    "type": "duplicate_sku_target",
                    "message": (
                        f"Found {len(dup_skus)} duplicate SKUs in target file. "
                        f"All rows with duplicate SKUs will be updated with the same source value."
                    ),
                    "skus": dup_skus,
                    "count": len(dup_skus),
                }
            )

    def _perform_comparison(self):
        """Execute join operations and column comparisons."""
        # Validate SKU columns exist
        if self.source_sku_col not in self.source_df.columns:
            raise ValueError(
                f"Source SKU column '{self.source_sku_col}' not found"
            )
        if self.target_sku_col not in self.target_df.columns:
            raise ValueError(
                f"Target SKU column '{self.target_sku_col}' not found"
            )

        # Rename SKU columns to common name and cast to Utf8 for consistent joining
        source_renamed = self.source_df.rename(
            {self.source_sku_col: "__SKU__"}
        ).with_columns(pl.col("__SKU__").cast(pl.Utf8))
        target_renamed = self.target_df.rename(
            {self.target_sku_col: "__SKU__"}
        ).with_columns(pl.col("__SKU__").cast(pl.Utf8))

        # Find common SKUs (inner join) and new SKUs (anti join)
        common_skus = source_renamed.select("__SKU__").join(
            target_renamed.select("__SKU__"), on="__SKU__", how="inner"
        )

        new_skus = source_renamed.select("__SKU__").join(
            target_renamed.select("__SKU__"), on="__SKU__", how="anti"
        )

        # Compare existing columns for common SKUs
        self._compare_common_rows(source_renamed, target_renamed, common_skus)

        # Process new columns to be added
        self._process_new_columns(source_renamed, target_renamed)

        # Prepare new rows for SKUs only in source
        self._prepare_additions(source_renamed, new_skus)

    def _compare_common_rows(
        self,
        source_df: pl.DataFrame,
        target_df: pl.DataFrame,
        common_skus: pl.DataFrame,
    ):
        """Compare values for common SKUs column by column."""
        update_mapping = self.column_mapping.get("updates", {})
        if not update_mapping:
            self.updates = []
            return

        # Filter to common SKUs only
        source_common = source_df.join(common_skus, on="__SKU__", how="inner")
        target_common = target_df.join(common_skus, on="__SKU__", how="inner")

        source_cols = list(update_mapping.keys())
        target_cols = list(update_mapping.values())

        target_rename = {col: f"{col}_target" for col in target_cols}

        src_selected = source_common.select(
            ["__SKU__"]
            + [c for c in source_cols if c in source_common.columns]
        )
        tgt_selected = target_common.select(
            ["__SKU__"]
            + [c for c in target_cols if c in target_common.columns]
        ).rename(target_rename)

        combined = src_selected.join(tgt_selected, on="__SKU__", how="inner")

        updates = []
        for src_col, tgt_col in update_mapping.items():
            tgt_renamed = f"{tgt_col}_target"
            if (
                src_col not in combined.columns
                or tgt_renamed not in combined.columns
            ):
                continue

            diffs = combined.filter(
                pl.col(src_col).is_not_null()
                & (
                    pl.col(tgt_renamed).is_null()
                    | (
                        pl.col(src_col).cast(pl.Utf8)
                        != pl.col(tgt_renamed).cast(pl.Utf8)
                    )
                )
            )

            if diffs.is_empty():
                continue

            col_updates = diffs.select(
                pl.col("__SKU__").alias("sku"),
                pl.lit(tgt_col).alias("column"),
                pl.lit(src_col).alias("source_column"),
                pl.col(tgt_renamed).alias("old_value"),
                pl.col(src_col).alias("new_value"),
            ).to_dicts()

            updates.extend(col_updates)

        self.updates = updates

    def _process_new_columns(
        self, source_df: pl.DataFrame, target_df: pl.DataFrame
    ):
        """Process new columns to be added to target file."""
        new_columns_mapping = self.column_mapping.get("new_columns", {})
        if not new_columns_mapping:
            return

        # Get target SKUs with row indices for alignment
        target_skus = target_df.select("__SKU__").with_row_index(
            name="row_idx"
        )

        existing_src_cols = [
            src
            for src in new_columns_mapping.keys()
            if src in source_df.columns
        ]
        if not existing_src_cols:
            return

        rename_map = {
            src: new_columns_mapping[src] for src in existing_src_cols
        }
        src_data = source_df.select(["__SKU__"] + existing_src_cols).rename(
            rename_map
        )

        aligned = target_skus.join(src_data, on="__SKU__", how="left")

        for new_col in rename_map.values():
            self.new_columns_data[new_col] = aligned.select(
                pl.col("row_idx"), pl.col(new_col)
            )

    def _prepare_additions(
        self, source_df: pl.DataFrame, new_skus: pl.DataFrame
    ):
        """Prepare new rows to be added for SKUs only in source.

        Args:
            source_df: Source DataFrame with renamed SKU column.
            new_skus: DataFrame of SKUs present only in source.
        """
        # Get source rows for new SKUs
        additions = source_df.join(new_skus, on="__SKU__", how="inner")

        # Combine all column mappings
        all_mappings = {}
        all_mappings.update(self.column_mapping.get("updates", {}))
        all_mappings.update(self.column_mapping.get("new_columns", {}))

        # Rename source columns to match target column names
        rename_map = {"__SKU__": self.target_sku_col}
        for source_col, target_col in all_mappings.items():
            if source_col in additions.columns:
                rename_map[source_col] = target_col

        self.additions = additions.rename(rename_map)

        # Keep only columns that exist in target (or will be added)
        mapped_target_cols = list(all_mappings.values()) + [
            self.target_sku_col
        ]
        cols_to_keep = [
            c for c in self.additions.columns if c in mapped_target_cols
        ]
        self.additions = self.additions.select(cols_to_keep)
