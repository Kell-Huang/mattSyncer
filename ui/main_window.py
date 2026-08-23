import os
import subprocess
import sys
import time
import traceback

import polars as pl
from openpyxl import load_workbook
from PySide6.QtCore import Qt, QThread, QTimer, Signal
from PySide6.QtGui import QColor, QIcon
from PySide6.QtWidgets import (
    QApplication,
    QFrame,
    QGraphicsDropShadowEffect,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QSplitter,
    QVBoxLayout,
    QWidget,
)

from core.differ import Differ
from core.writer import Writer
from ui.widgets.action_panel import ActionPanel
from ui.widgets.file_panel import FilePanel
from ui.widgets.formula_panel import FormulaPanel
from ui.widgets.info_panel import InfoPanel
from ui.widgets.mapping_panel import MappingPanel
from ui.widgets.sku_panel import SkuPanel


# Global exception hook to surface any import / runtime errors
def global_exception_handler(exc_type, exc_value, exc_traceback):
    traceback.print_exception(exc_type, exc_value, exc_traceback)
    sys.__excepthook__(exc_type, exc_value, exc_traceback)


sys.excepthook = global_exception_handler


class ExecuteWorker(QThread):
    progress_updated = Signal(str, int)
    step_completed = Signal(str, float)
    finished = Signal(dict)
    error = Signal(str)
    writer_progress = Signal(str)

    def __init__(self, data):
        super().__init__()
        self.data = data

    def run(self):
        t0 = time.time()
        try:
            source_files = self.data.get("source_files", [])
            if not source_files:
                self.error.emit("No source files provided.")
                return

            # Base session (target-related fields remain constant)
            base_session = {
                "target_path": self.data["target_path"],
                "source_sku_col": self.data["source_sku_col"],
                "target_sku_col": self.data["target_sku_col"],
                "column_mapping": self.data["column_mapping"],
                "output_path": self.data.get("output_path", ""),
            }
            formula_columns = self.data.get("formula_columns")
            if formula_columns:
                base_session["formula_columns"] = formula_columns

            # Flatten all files/sheets into a single list of virtual source units
            flat_sources = []
            for file_info in source_files:
                path = file_info["path"]
                if path.endswith(".csv"):
                    flat_sources.append({"path": path, "sheet_name": None})
                else:
                    wb = load_workbook(path, read_only=True)
                    flat_sources.extend(
                        {"path": path, "sheet_name": sname}
                        for sname in wb.sheetnames
                    )
                    wb.close()

            if not flat_sources:
                self.error.emit("No valid source sheets found.")
                return

            total_updates = 0
            total_new_skus = 0
            intermediate_df = None
            first_valid_processed = False

            # Process each virtual source file sequentially
            for idx, src in enumerate(flat_sources, start=1):
                session = dict(base_session)
                session["source_path"] = src["path"]
                sheet_name = src.get("sheet_name")
                session["sheet_name"] = sheet_name

                # Display name for logging
                if sheet_name:
                    display_name = (
                        f"{os.path.basename(src['path'])} - {sheet_name}"
                    )
                else:
                    display_name = os.path.basename(src["path"])
                self.writer_progress.emit(f"Processing: {display_name}")

                # Decide whether to use the target on disk or the accumulated intermediate_df
                if not first_valid_processed:
                    # First valid source: read target from disk
                    differ = Differ(session, sheet_name=sheet_name)
                else:
                    # Subsequent sources: use the intermediate DataFrame as target
                    differ = Differ(
                        session,
                        target_df=intermediate_df,
                        sheet_name=sheet_name,
                    )

                diff = differ.compare()

                # Check target duplicates only once, on the first valid source
                if not first_valid_processed:
                    differ.warn_target_duplicates()

                # Print warnings except missing_sku (those are handled separately)
                for w in diff.get("warnings", []):
                    if w.get("type") != "missing_sku":
                        print("[WARNING]", w.get("message", str(w)))

                # If this sheet was skipped due to missing SKU column, warn and continue
                if any(
                    w["type"] == "missing_sku"
                    for w in diff.get("warnings", [])
                ):
                    self.writer_progress.emit(
                        f"Skipped: {display_name} (missing SKU column)"
                    )
                    self.step_completed.emit(f"Source {idx} skipped", 0)
                    continue

                # This sheet is valid – accumulate statistics
                total_updates += len(diff["updates"])
                total_new_skus += diff["new_skus"]

                if not first_valid_processed:
                    # First valid source: create intermediate_df from scratch
                    writer = Writer(session, diff)
                    source_lf = writer._get_lazy_source()
                    source_lf = writer._apply_updates(source_lf)

                    # Merge new columns (first valid source)
                    new_columns_mapping = writer.column_mapping.get(
                        "new_columns", {}
                    )
                    if (
                        new_columns_mapping
                        and writer.new_columns_combined is not None
                    ):
                        new_cols_lf = writer.new_columns_combined.lazy()
                        source_lf = source_lf.with_row_index(name="_row_idx_")
                        source_lf = source_lf.join(
                            new_cols_lf.rename({"row_idx": "_row_idx_"}),
                            on="_row_idx_",
                            how="left",
                        )
                        source_lf = source_lf.drop("_row_idx_")

                    # Merge additions (new SKUs)
                    if (
                        writer.additions is not None
                        and len(writer.additions) > 0
                    ):
                        add_lf = writer.additions.lazy()
                        source_schema = source_lf.collect_schema()
                        source_cols = source_schema.names()
                        add_schema = add_lf.collect_schema()
                        for col in source_cols:
                            if col not in add_schema.names():
                                add_lf = add_lf.with_columns(
                                    pl.lit(None).alias(col)
                                )
                        add_lf = add_lf.select(source_cols)
                        add_lf = add_lf.select(
                            pl.all().cast(pl.Utf8, strict=False)
                        )
                        source_lf = source_lf.select(
                            pl.all().cast(pl.Utf8, strict=False)
                        )
                        source_lf = pl.concat([source_lf, add_lf])
                        source_lf = source_lf.select(
                            pl.all().cast(pl.Utf8, strict=False)
                        )

                    intermediate_df = source_lf.collect()
                    first_valid_processed = True
                    self.step_completed.emit(
                        f"Source {idx} loaded", time.time() - t0
                    )
                else:
                    # Subsequent valid source: incremental update on intermediate_df
                    writer = Writer(session, diff, result_df=intermediate_df)
                    source_lf = writer._get_lazy_source()
                    source_lf = writer._apply_updates(source_lf)

                    # Only add columns that do not already exist in the intermediate result
                    if writer.new_columns_combined is not None:
                        existing_columns = intermediate_df.columns
                        new_cols_df = writer.new_columns_combined
                        truly_new_cols = [
                            c
                            for c in new_cols_df.columns
                            if c != "row_idx" and c not in existing_columns
                        ]
                        if truly_new_cols:
                            new_cols_lf = new_cols_df.select(
                                ["row_idx"] + truly_new_cols
                            ).lazy()
                            source_lf = source_lf.with_row_index(
                                name="_row_idx_"
                            )
                            source_lf = source_lf.join(
                                new_cols_lf.rename({"row_idx": "_row_idx_"}),
                                on="_row_idx_",
                                how="left",
                            )
                            source_lf = source_lf.drop("_row_idx_")

                    # Merge additions
                    if (
                        writer.additions is not None
                        and len(writer.additions) > 0
                    ):
                        add_lf = writer.additions.lazy()
                        source_schema = source_lf.collect_schema()
                        source_cols = source_schema.names()
                        add_schema = add_lf.collect_schema()
                        for col in source_cols:
                            if col not in add_schema.names():
                                add_lf = add_lf.with_columns(
                                    pl.lit(None).alias(col)
                                )
                        add_lf = add_lf.select(source_cols)
                        add_lf = add_lf.select(
                            pl.all().cast(pl.Utf8, strict=False)
                        )
                        source_lf = source_lf.select(
                            pl.all().cast(pl.Utf8, strict=False)
                        )
                        source_lf = pl.concat([source_lf, add_lf])
                        source_lf = source_lf.select(
                            pl.all().cast(pl.Utf8, strict=False)
                        )

                    intermediate_df = source_lf.collect()
                    self.step_completed.emit(
                        f"Source {idx} loaded", time.time() - t0
                    )

                # Update progress
                progress = 5 + int(
                    idx / len(flat_sources) * 55
                )  # 5% to 60% during processing
                self.progress_updated.emit(
                    f"Processing file {idx}/{len(flat_sources)}", progress
                )

            # If no valid source was found, abort
            if not first_valid_processed:
                self.error.emit(
                    "No valid source data found. All sheets were skipped because they lack the required SKU column."
                )
                return

            # ---------- Final write ----------
            self.progress_updated.emit("Writing...", 60)
            empty_diff = {
                "updates": [],
                "additions": None,
                "new_columns_data": {},
            }
            final_writer = Writer(
                base_session, empty_diff, result_df=intermediate_df
            )
            final_writer.progress_updated.connect(self.writer_progress.emit)
            final_writer.write()

            self.step_completed.emit("File written", time.time() - t0)
            self.progress_updated.emit("Report...", 90)
            report = {
                "cells_updated": total_updates,
                "new_skus": total_new_skus,
                "elapsed_time": time.time() - t0,
            }
            self.progress_updated.emit("Complete", 100)
            self.finished.emit(report)

        except Exception as e:
            import traceback

            self.error.emit(f"{e}\n{traceback.format_exc()}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.session_data = {}
        self.worker = None
        self.init_ui()
        self.connect_signals()
        self._initial_sizes_set = False

    def init_ui(self):
        # Set window icon (icon file located in the same directory as this file)
        icon_path = os.path.join(
            os.path.dirname(__file__), "dataSyncerIcon.ico"
        )
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.setWindowTitle("MATTUpdater - Source to Target File Update Tool")
        self.setMinimumSize(1100, 750)
        self.resize(1250, 820)

        # Load QSS stylesheet
        qss_path = os.path.join(os.path.dirname(__file__), "style.qss")
        if os.path.exists(qss_path):
            with open(qss_path, "r") as f:
                self.setStyleSheet(f.read())

        central = QWidget()
        central.setObjectName("centralWidget")
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Header
        header = QFrame()
        header.setObjectName("headerFrame")
        header.setFixedHeight(56)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(14, 6, 14, 6)
        header_layout.setSpacing(10)

        title_container = QWidget()
        title_layout = QVBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 0)
        title_layout.setSpacing(0)
        title_layout.addWidget(QLabel("MATTUpdater", objectName="titleLabel"))
        title_layout.addWidget(
            QLabel(
                "Source to Target File Update Tool", objectName="subtitleLabel"
            )
        )
        header_layout.addWidget(title_container)
        header_layout.addStretch()
        header_layout.addWidget(QLabel("v1.0.0", objectName="versionLabel"))
        main_layout.addWidget(header)

        # Content
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(12, 8, 12, 8)
        content_layout.setSpacing(8)

        # Create card containers for all panels (component creation order remains unchanged)
        # LogPanel has been removed completely from the UI.
        self.file_panel_container = self._create_card_container(
            FilePanel(), "fileCard", "File Panel"
        )
        self.sku_panel_container = self._create_card_container(
            SkuPanel(), "skuCard", "SKU Panel"
        )
        self.formula_panel_container = self._create_card_container(
            FormulaPanel(), "formulaCard", "Formula Panel"
        )
        self.mapping_panel_container = self._create_card_container(
            MappingPanel(), "mappingCard", "Mapping Panel"
        )
        self.action_panel_container = self._create_card_container(
            ActionPanel(), "actionCard", "Action Panel", show_title=False
        )
        self.info_panel_container = self._create_card_container(
            InfoPanel(), "infoCard", "Info Panel", show_title=False
        )

        # Top status bar: original Info Panel, fixed height 35px.
        self.info_panel_container.setFixedHeight(35)
        # Remove extra padding for the slim strip
        info_layout = self.info_panel_container.layout()
        info_layout.setContentsMargins(8, 2, 8, 2)
        info_layout.setSpacing(0)
        content_layout.addWidget(self.info_panel_container)

        # Main horizontal splitter: left configuration area (~30%) | right mapping area (~70%)
        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setHandleWidth(2)

        # Left configuration area: File, SKU, Formula stacked vertically.
        # They must NOT stretch vertically; they should fit their content and leave blank space below.
        left_config = QWidget()
        left_config_layout = QVBoxLayout(left_config)
        left_config_layout.setContentsMargins(0, 0, 0, 0)
        left_config_layout.setSpacing(8)
        left_config_layout.addWidget(self.file_panel_container)
        left_config_layout.addWidget(self.sku_panel_container)
        left_config_layout.addWidget(self.formula_panel_container)
        left_config_layout.addStretch(
            1
        )  # Absorb extra vertical space, keep panels compact

        self.main_splitter.addWidget(left_config)
        self.main_splitter.addWidget(self.mapping_panel_container)
        self.main_splitter.setStretchFactor(0, 30)  # left config ~30%
        self.main_splitter.setStretchFactor(1, 70)  # mapping ~70%

        content_layout.addWidget(self.main_splitter, stretch=1)

        # Bottom Action panel only. LogPanel has been removed.
        self.action_panel_container.setFixedHeight(60)
        # Remove extra padding for the action strip
        action_layout = self.action_panel_container.layout()
        action_layout.setContentsMargins(8, 4, 8, 4)
        action_layout.setSpacing(0)
        content_layout.addWidget(self.action_panel_container)

        main_layout.addWidget(content, stretch=1)

        # Status bar (hidden to preserve component)
        status_bar = QFrame()
        status_bar.setObjectName("statusBarFrame")
        status_bar.setFixedHeight(0)
        status_bar.hide()
        status_layout = QHBoxLayout(status_bar)
        status_layout.setContentsMargins(24, 2, 24, 2)
        status_layout.setSpacing(10)
        self.status_label = QLabel("Ready")
        self.status_label.setObjectName("statusLabel")
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        main_layout.addWidget(status_bar)

        # Extract actual panel instances from containers for signal connections
        self.file_panel = self.file_panel_container.findChild(FilePanel)
        self.sku_panel = self.sku_panel_container.findChild(SkuPanel)
        self.formula_panel = self.formula_panel_container.findChild(
            FormulaPanel
        )
        self.mapping_panel = self.mapping_panel_container.findChild(
            MappingPanel
        )
        self.action_panel = self.action_panel_container.findChild(ActionPanel)
        self.info_panel = self.info_panel_container.findChild(InfoPanel)

    def _create_card_container(
        self,
        widget: QWidget,
        object_name: str,
        title: str,
        show_title: bool = True,
    ) -> QFrame:
        """
        Wrap a widget in a card-style QFrame container with padding, border,
        drop shadow, and a panel title.

        Args:
            widget: The panel widget to embed.
            object_name: QSS object name for the card frame.
            title: Panel title text.
            show_title: If False, the title label is hidden (but still exists).
        """
        card = QFrame()
        card.setProperty("card", "true")
        card.setObjectName(object_name)

        # Container layout
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        # Panel title
        title_label = QLabel(title)
        title_label.setProperty("panelTitle", "true")
        if not show_title:
            title_label.hide()
        layout.addWidget(title_label)

        # Actual widget (stretch to fill remaining space)
        layout.addWidget(widget, 1)

        # Apply drop shadow effect
        shadow = QGraphicsDropShadowEffect(card)
        shadow.setBlurRadius(12)
        shadow.setColor(QColor(0, 0, 0, 10))  # alpha ~ 0.04
        shadow.setOffset(0, 2)
        card.setGraphicsEffect(shadow)

        return card

    def _set_initial_sizes(self):
        """
        Set initial splitter sizes based on current window dimensions and
        the desired proportions (from design specification).
        """
        # Available content width after 12px margins on each side.
        content_width = max(800, self.width() - 24)

        # Main horizontal splitter: left config (~30%) | right mapping (~70%).
        # The handle width is subtracted to avoid rounding overflows.
        left_width = int(content_width * 0.30)
        right_width = (
            content_width - left_width - self.main_splitter.handleWidth()
        )
        self.main_splitter.setSizes([left_width, right_width])

    def showEvent(self, event):
        """Set initial splitter sizes after the window is fully shown."""
        super().showEvent(event)
        if not self._initial_sizes_set:
            QTimer.singleShot(0, self._apply_initial_sizes)
            self._initial_sizes_set = True

    def _apply_initial_sizes(self):
        """Apply initial sizes and then print current sizes for debugging."""
        self._set_initial_sizes()

    # ---- Signal connections ----
    def connect_signals(self):
        self.file_panel.files_selected.connect(self.on_files_selected)
        self.sku_panel.sku_confirmed.connect(self.on_sku_confirmed)
        # Log-related signals are now redirected to console print statements.
        self.sku_panel.detection_warning.connect(
            lambda message: print("[WARNING]", message)
        )
        self.action_panel.preview_clicked.connect(self.on_preview)
        self.action_panel.execute_clicked.connect(self.on_execute)
        self.mapping_panel.error_occurred.connect(
            lambda message: print("[ERROR]", message)
        )

    # ---- Business logic methods ----
    def on_files_selected(self, sp, tp):
        self.status_label.setText("Loading...")
        QApplication.processEvents()
        self.info_panel.update_source_info(sp)
        self.info_panel.update_target_info(tp)
        self.sku_panel.detect_columns(sp, tp)
        print("[INFO]", "Files loaded")
        self.status_label.setText("Confirm SKU columns")

    def on_sku_confirmed(self, ssk, tsk):
        self.status_label.setText("Matching...")
        QApplication.processEvents()
        self.mapping_panel.perform_matching(
            self.file_panel.get_source_files(),
            self.file_panel.target_path,
            ssk,
            tsk,
        )
        print("[INFO]", f"SKU: {ssk} -> {tsk}")
        self.status_label.setText("Ready")
        # Update formula panel with target column names
        self.formula_panel.set_columns(self.mapping_panel.get_target_columns())

    def on_preview(self):
        if not self._validate():
            return
        m = self.mapping_panel.get_confirmed_mapping()
        ssk, tsk = self.sku_panel.get_sku_columns()
        formula_cols = self.formula_panel.get_selected_columns()
        try:
            print("[INFO]", "Preview...")
            data = {
                "source_path": self.file_panel.source_path,
                "source_files": self.file_panel.get_source_files(),
                "target_path": self.file_panel.target_path,
                "source_sku_col": ssk,
                "target_sku_col": tsk,
                "column_mapping": m,
            }
            if formula_cols:
                data["formula_columns"] = formula_cols
            diff = Differ(data).compare()
            print(
                "[INFO]",
                f"Updates: {len(diff.get('updates',[]))}, "
                f"New SKUs: {diff.get('new_skus',0)}, "
                f"New cols: {diff.get('new_columns_count',0)}",
            )
            for w in diff.get("warnings", []):
                print("[WARNING]", w.get("message", str(w)))
        except Exception as e:
            print("[ERROR]", str(e))

    def on_execute(self):
        if not self._validate():
            return
        m = self.mapping_panel.get_confirmed_mapping()
        ssk, tsk = self.sku_panel.get_sku_columns()
        formula_cols = self.formula_panel.get_selected_columns()
        self.session_data = {
            "source_files": self.file_panel.get_source_files(),
            "target_path": self.file_panel.target_path,
            "source_sku_col": ssk,
            "target_sku_col": tsk,
            "column_mapping": m,
            "output_path": self.file_panel.get_output_path(),
        }
        if formula_cols:
            self.session_data["formula_columns"] = formula_cols

        print("[INFO]", "Starting...")
        self.action_panel.set_running(True)
        self.info_panel.status.set_running()
        self.worker = ExecuteWorker(self.session_data)
        self.worker.progress_updated.connect(self.action_panel.set_progress)
        self.worker.step_completed.connect(
            lambda step_name, duration: print(
                "[STEP]", f"{step_name:<35} {duration:.1f}s"
            )
        )
        self.worker.writer_progress.connect(
            lambda message: print("[INFO]", message)
        )
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)
        self.worker.start()

    def on_finished(self, report):
        self.action_panel.set_running(False)
        self.info_panel.status.set_done()
        e = report.get("elapsed_time", 0)
        m, s = int(e // 60), int(e % 60)
        print("[INFO]", f"Done in {m}m{s}s")

        summary = (
            f"Updated: {report.get('cells_updated', 0):,}\n"
            f"New SKUs: {report.get('new_skus', 0):,}\n"
            f"Elapsed: {m}m {s}s"
        )
        print("[INFO]", summary.replace("\n", " | "))

        out = self.session_data.get("output_path", "")
        msg = f"{summary}\n\nOpen folder?\n{out}"
        if (
            QMessageBox.question(
                self, "Done", msg, QMessageBox.Yes | QMessageBox.No
            )
            == QMessageBox.Yes
        ):
            folder = os.path.dirname(out)
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.run(["open", folder])
            else:
                subprocess.run(["xdg-open", folder])

    def on_error(self, msg):
        self.action_panel.set_running(False)
        # Extract first non-empty line as a concise summary for UI
        summary = next(
            (line.strip() for line in msg.splitlines() if line.strip()),
            "Unknown error",
        )
        self.info_panel.status.set_error(summary)
        print("[ERROR]", msg)  # Keep full traceback in console
        QMessageBox.critical(
            self,
            "Error",
            f"{summary}\n\nPlease check the console output for details.",
        )

    def _validate(self):
        if not self.file_panel.validate():
            QMessageBox.warning(self, "Error", "Select files.")
            return False
        if not self.sku_panel.validate():
            QMessageBox.warning(self, "Error", "Confirm SKU columns.")
            return False
        if not self.mapping_panel.validate():
            return False
        return True
