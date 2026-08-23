import os
from functools import partial

import polars as pl
from openpyxl import load_workbook
from PySide6.QtCore import QEvent, Qt, Signal
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from core.column_matcher import ColumnMatcher
from utils.encoding import detect_encoding


class FloatingDropdown(QListWidget):
    """Floating dropdown list for column search."""

    item_selected = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName(
            "floatingDropdown"
        )  # Style handled by QSS via parent
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_ShowWithoutActivating)
        self.setFocusPolicy(Qt.NoFocus)
        self.setMinimumHeight(100)
        self.setMaximumHeight(180)
        self.itemClicked.connect(self._on_item_clicked)
        self.hide()

    def set_items(self, items):
        self.clear()
        for text in items:
            item = QListWidgetItem(text)
            self.addItem(item)

    def filter_items(self, text):
        for i in range(self.count()):
            item = self.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def show_at(self, global_pos, width):
        self.setFixedWidth(width)
        if self.parent() is None:
            self.move(global_pos)
        else:
            parent_pos = self.parent().mapFromGlobal(global_pos)
            self.move(parent_pos)
        self.show()
        self.raise_()

    def _on_item_clicked(self, item):
        self.item_selected.emit(item.text())
        self.hide()


class SearchableCell(QWidget):
    """Unified cell widget for all Target Column types."""

    selection_made = Signal(str)

    def __init__(self, all_items, initial_text="", editable=True, parent=None):
        super().__init__(parent)
        self.all_items = all_items

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.edit = QLineEdit()
        self.edit.setObjectName("tableSearchEdit")  # Style handled by QSS
        self.edit.setFixedHeight(30)  # Keep consistent row height
        self.edit.setText(initial_text)
        self.edit.setReadOnly(not editable)
        if editable:
            self.edit.setPlaceholderText("Type to search...")
            self.edit.textChanged.connect(self._on_text_changed)
            self.edit.installEventFilter(self)
        layout.addWidget(self.edit)

        self._dropdown = None

    def _get_dropdown(self):
        if self._dropdown is None:
            # Use the main window as parent so the dropdown inherits the QSS stylesheet
            self._dropdown = FloatingDropdown(self.window())
            self._dropdown.set_items(self.all_items)
            self._dropdown.item_selected.connect(self._on_item_selected)
        return self._dropdown

    def _on_text_changed(self, text):
        dropdown = self._get_dropdown()
        if not text.strip():
            dropdown.hide()
            return
        dropdown.filter_items(text)
        global_pos = self.edit.mapToGlobal(self.edit.rect().bottomLeft())
        dropdown.show_at(global_pos, self.edit.width())

    def _on_item_selected(self, text):
        self.edit.blockSignals(True)
        self.edit.setText(text)
        self.edit.blockSignals(False)
        if self._dropdown:
            self._dropdown.hide()
        self.selection_made.emit(text)
        self.edit.setFocus()

    def eventFilter(self, obj, event):
        if obj == self.edit:
            if event.type() == QEvent.FocusOut:
                if self._dropdown and self._dropdown.isVisible():
                    self._dropdown.hide()
            elif event.type() == QEvent.KeyPress:
                key_event = event
                dropdown = self._get_dropdown() if self._dropdown else None
                if key_event.key() == Qt.Key_Escape:
                    if dropdown:
                        dropdown.hide()
                    return True
                if (
                    key_event.key() == Qt.Key_Down
                    and dropdown
                    and dropdown.isVisible()
                ):
                    if dropdown.count() > 0:
                        dropdown.setCurrentRow(0)
                    return True
                if key_event.key() in (Qt.Key_Return, Qt.Key_Enter):
                    if (
                        dropdown
                        and dropdown.isVisible()
                        and dropdown.currentItem()
                    ):
                        self._on_item_selected(dropdown.currentItem().text())
                        return True
        return super().eventFilter(obj, event)

    def get_selected_text(self):
        return self.edit.text().strip()


class MappingPanel(QWidget):
    """Column mapping panel with searchable target columns and batch selection."""

    mapping_confirmed = Signal(dict)
    error_occurred = Signal(str)

    def __init__(self):
        super().__init__()
        self.column_matcher = ColumnMatcher()
        self.mapping_results = []
        self.prefix_info = {}
        self.new_columns = {}
        self.target_columns = []
        self.last_checked_row = -1
        self._column_cache = {}
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(
            4
        )  # Tighten vertical spacing inside mapping panel

        # Summary label
        self.summary_label = QLabel(
            "Matched: 0/0  |  Selected: 0  |  New: 0  |  Unmatched: 0"
        )
        self.summary_label.setObjectName(
            "summaryLabel"
        )  # Style handled by QSS
        main_layout.addWidget(self.summary_label)

        # Prefix label
        self.prefix_label = QLabel()
        self.prefix_label.setObjectName("prefixLabel")  # Style handled by QSS
        self.prefix_label.hide()
        main_layout.addWidget(self.prefix_label)

        # Toolbar row
        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        self.select_all_btn = QPushButton("Deselect All")
        self.select_all_btn.setProperty("secondary", "true")
        self.select_all_btn.setFixedWidth(100)
        self.select_all_btn.clicked.connect(self.toggle_select_all)

        btn_dir = QPushButton("Direct Only")
        btn_dir.setProperty("secondary", "true")
        btn_dir.setFixedWidth(90)
        btn_dir.clicked.connect(lambda: self.filter_by_type("direct"))

        btn_unm = QPushButton("Unmatched Only")
        btn_unm.setProperty("secondary", "true")
        btn_unm.setFixedWidth(110)
        btn_unm.clicked.connect(lambda: self.filter_by_type("none"))

        self.search = QLineEdit()
        self.search.setObjectName("searchInput")
        self.search.setPlaceholderText("Search...")
        self.search.setFixedHeight(26)
        self.search.setFixedWidth(180)
        self.search.textChanged.connect(self.filter_table)

        toolbar.addWidget(self.select_all_btn)
        toolbar.addWidget(btn_dir)
        toolbar.addWidget(btn_unm)
        toolbar.addStretch()
        toolbar.addWidget(self.search)
        main_layout.addLayout(toolbar)

        # Table: must stretch to fill all remaining vertical space.
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ["Sel/Type", "Source Column", "", "Target Column", "Act"]
        )
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setAlternatingRowColors(True)

        vh = self.table.verticalHeader()
        vh.setDefaultAlignment(Qt.AlignCenter)

        h = self.table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Fixed)
        h.setSectionResizeMode(1, QHeaderView.Stretch)
        h.setSectionResizeMode(2, QHeaderView.Fixed)
        h.setSectionResizeMode(3, QHeaderView.Stretch)
        h.setSectionResizeMode(4, QHeaderView.Fixed)
        self.table.setColumnWidth(0, 120)
        self.table.setColumnWidth(2, 18)
        self.table.setColumnWidth(4, 60)

        main_layout.addWidget(self.table, 1)

    # ================= Business logic (unchanged) =================
    def perform_matching(self, source_files, target_path, ssk, tsk):
        """Perform column matching using columns from all source files/sheets
        and the target file (headers only, no data loaded).
        """
        try:
            all_source_cols = []
            seen_source_cols = set()
            for file_info in source_files:
                file_path = file_info["path"]
                columns = self._get_file_columns(file_path)
                for col_name in columns:
                    if col_name not in seen_source_cols:
                        seen_source_cols.add(col_name)
                        all_source_cols.append(col_name)

            src_cols = [c for c in all_source_cols if c != ssk]
            self.target_columns = self._get_file_columns(target_path)
            self.mapping_results, self.prefix_info = (
                self.column_matcher.match_columns(
                    src_cols, self.target_columns
                )
            )

            type_order = {"direct": 0, "indirect": 1, "none": 2}
            self.mapping_results.sort(
                key=lambda item: type_order.get(item["match_type"], 2)
            )

            pref = self.prefix_info.get("most_common_prefix", "")
            self.prefix_label.setText(
                f"Detected prefix: '{pref}' - New columns: {pref}[source]"
                if pref
                else "No common prefix. New columns will use source names directly."
            )
            self.prefix_label.show()

            self.new_columns = {}
            self.last_checked_row = -1
            for r in self.mapping_results:
                r["selected"] = True
                r["original_match_type"] = r["match_type"]
            self.populate_table()
            self.update_summary()

        except Exception as e:
            self.error_occurred.emit(f"Matching failed: {e}")

    def populate_table(self):
        self.table.setRowCount(len(self.mapping_results))
        all_targets = self.target_columns

        for row, res in enumerate(self.mapping_results):
            cw = QWidget()
            chl = QHBoxLayout(cw)
            chl.setContentsMargins(4, 0, 2, 0)
            chl.setSpacing(4)
            chl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

            cb = QCheckBox()
            cb.setFocusPolicy(Qt.StrongFocus)
            cb.clicked.connect(partial(self._on_checkbox_clicked, row))
            cb.setChecked(res.get("selected", False))

            # Badge with type property for QSS styling
            type_text, badge_type = self._type_info(res)
            badge = QLabel(type_text)
            badge.setAlignment(Qt.AlignCenter)
            badge.setProperty("type", badge_type)  # Used by QSS
            badge.setMinimumWidth(
                70
            )  # Ensure "INDIRECT" and "UNMATCHED" are fully visible
            badge.setFixedHeight(18)
            chl.addWidget(cb)
            chl.addWidget(badge)
            self.table.setCellWidget(row, 0, cw)

            si = QTableWidgetItem(res["source"])
            si.setFlags(si.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 1, si)

            ai = QTableWidgetItem("→")
            ai.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 2, ai)

            is_unmatched = (
                res["match_type"] == "none"
                and res["source"] not in self.new_columns
            )
            is_indirect = res["match_type"] == "indirect"
            is_new = res["source"] in self.new_columns

            if is_new:
                search_cell = SearchableCell(
                    all_targets,
                    self.new_columns[res["source"]],
                    editable=False,
                )
                self.table.setCellWidget(row, 3, search_cell)
            elif is_unmatched or is_indirect:
                initial = (
                    res["target"] if is_indirect and res["target"] else ""
                )
                search_cell = SearchableCell(
                    all_targets, initial, editable=True
                )
                search_cell.selection_made.connect(
                    lambda text, r=row: self._on_search_select(r, text)
                )
                self.table.setCellWidget(row, 3, search_cell)
            else:
                search_cell = SearchableCell(
                    all_targets, res["target"] or "", editable=False
                )
                self.table.setCellWidget(row, 3, search_cell)

            self._set_action_cell(row, res)

    def _type_info(self, res):
        if res["source"] in self.new_columns:
            return ("NEW", "new")
        if res["match_type"] == "direct":
            return ("DIRECT", "direct")
        elif res["match_type"] == "indirect":
            return ("INDIRECT", "indirect")
        else:
            return ("UNMATCHED", "unmatched")

    def _set_action_cell(self, row, res):
        if res["source"] in self.new_columns:
            btn = QPushButton("Del")
            btn.setProperty("secondary", "true")
            btn.setProperty("small", "true")
            btn.setFixedSize(50, 24)
            btn.clicked.connect(
                lambda checked, r=row: self.remove_new_column(r)
            )
            self.table.setCellWidget(row, 4, btn)
        elif (
            res["match_type"] == "none"
            and res["source"] not in self.new_columns
        ) or (
            res["match_type"] == "indirect"
            and res["source"] not in self.new_columns
        ):
            btn = QPushButton("Add")
            btn.setProperty("secondary", "true")
            btn.setProperty("small", "true")
            btn.setFixedSize(50, 24)
            btn.clicked.connect(
                lambda checked, r=row: self.add_single_new_column(r)
            )
            self.table.setCellWidget(row, 4, btn)
        else:
            self.table.setItem(row, 4, QTableWidgetItem(""))

    def _on_search_select(self, row, text):
        if row < len(self.mapping_results):
            res = self.mapping_results[row]
            if not text.strip():
                if res["match_type"] == "indirect":
                    res["match_type"] = "none"
                    res["target"] = None
                    self.refresh_table()
                    return
                else:
                    res["target"] = None
            else:
                res["target"] = text

    def _get_checkbox(self, row):
        cw = self.table.cellWidget(row, 0)
        if cw:
            return cw.findChild(QCheckBox)
        return None

    def _sync_all_checkboxes(self):
        for row in range(self.table.rowCount()):
            cb = self._get_checkbox(row)
            if cb:
                cb.blockSignals(True)
                cb.setChecked(self.mapping_results[row].get("selected", False))
                cb.blockSignals(False)

    def _on_checkbox_clicked(self, row):
        shift_held = QApplication.keyboardModifiers() == Qt.ShiftModifier
        if shift_held and self.last_checked_row >= 0:
            start = min(self.last_checked_row, row)
            end = max(self.last_checked_row, row)
            for r in range(start, end + 1):
                self.mapping_results[r]["selected"] = True
            self._sync_all_checkboxes()
        else:
            new_state = not self.mapping_results[row].get("selected", False)
            self.mapping_results[row]["selected"] = new_state
            cb = self._get_checkbox(row)
            if cb:
                cb.setChecked(new_state)

        self.last_checked_row = row
        self.update_summary()
        self._update_select_all_btn()

    def add_single_new_column(self, row):
        src = self.mapping_results[row]["source"]
        pref = self.prefix_info.get("most_common_prefix", "")
        new_name = self.column_matcher.generate_new_column_name(
            src, pref, self.target_columns
        )
        if new_name is None:
            QMessageBox.warning(
                self, "Conflict", f"Column '{pref}{src}' already exists."
            )
            return
        self.new_columns[src] = new_name
        if "original_match_type" not in self.mapping_results[row]:
            self.mapping_results[row]["original_match_type"] = (
                self.mapping_results[row]["match_type"]
            )
        self.mapping_results[row]["match_type"] = "new"
        self.mapping_results[row]["target"] = new_name
        self.mapping_results[row]["selected"] = True
        self.refresh_table()
        self.update_summary()

    def remove_new_column(self, row):
        src = self.mapping_results[row]["source"]
        if src in self.new_columns:
            del self.new_columns[src]
            original_type = self.mapping_results[row].get(
                "original_match_type", "none"
            )
            self.mapping_results[row]["match_type"] = original_type
            self.mapping_results[row].pop("original_match_type", None)
            self.mapping_results[row]["target"] = None
            self.mapping_results[row]["selected"] = False
        self.refresh_table()
        self.update_summary()

    def refresh_table(self):
        self.table.setRowCount(0)
        self.populate_table()

    def toggle_select_all(self):
        if all(r.get("selected", False) for r in self.mapping_results):
            self.deselect_all()
        else:
            self.select_all()

    def select_all(self):
        for r in self.mapping_results:
            r["selected"] = True
        self.refresh_table()
        self.update_summary()
        self._update_select_all_btn()

    def deselect_all(self):
        for r in self.mapping_results:
            r["selected"] = False
        self.refresh_table()
        self.update_summary()
        self._update_select_all_btn()

    def filter_by_type(self, mt):
        for r in self.mapping_results:
            r["selected"] = r["match_type"] == mt
        self.refresh_table()
        self.update_summary()
        self._update_select_all_btn()

    def _update_select_all_btn(self):
        if all(r.get("selected", False) for r in self.mapping_results):
            self.select_all_btn.setText("Deselect All")
        else:
            self.select_all_btn.setText("Select All")

    def filter_table(self, text):
        t = text.lower()
        for row in range(self.table.rowCount()):
            src_item = self.table.item(row, 1)
            if not src_item:
                continue
            src = src_item.text().lower()
            tw = self.table.cellWidget(row, 3)
            if isinstance(tw, SearchableCell):
                tgt = tw.get_selected_text().lower()
            else:
                tgt = ""
            self.table.setRowHidden(row, not (t in src or t in tgt))

    def update_summary(self):
        total = len(self.mapping_results)
        mat = sum(1 for r in self.mapping_results if r["match_type"] != "none")
        sel = sum(1 for r in self.mapping_results if r.get("selected") is True)
        unm = total - mat
        new = len(self.new_columns)
        self.summary_label.setText(
            f"Matched: {mat}/{total}  |  Selected: {sel}  |  New: {new}  |  Unmatched: {unm}"
        )

    def get_confirmed_mapping(self):
        m = {
            "updates": {},
            "new_columns": {},
            "prefix": self.prefix_info.get("most_common_prefix", ""),
        }
        for r in self.mapping_results:
            if r.get("selected"):
                s, t = r["source"], r["target"]
                if s in self.new_columns and t:
                    m["new_columns"][s] = t
                elif r["match_type"] != "none" and t:
                    m["updates"][s] = t
        return m

    def validate(self):
        us = [
            r["source"]
            for r in self.mapping_results
            if r["match_type"] == "none"
            and r.get("selected")
            and r["source"] not in self.new_columns
        ]
        if us:
            if (
                QMessageBox.warning(
                    self,
                    "Unmatched",
                    f"Selected but unmatched: {', '.join(us)}\nContinue?",
                    QMessageBox.Ok | QMessageBox.Ignore,
                    QMessageBox.Ok,
                )
                == QMessageBox.Ok
            ):
                return False
        if sum(1 for r in self.mapping_results if r.get("selected")) == 0:
            QMessageBox.warning(self, "None", "Select at least one column.")
            return False
        self.mapping_confirmed.emit(self.get_confirmed_mapping())
        return True

    def get_target_columns(self) -> list:
        return self.target_columns

    def _get_file_columns(self, file_path: str) -> list:
        """Return column names for the given file, using cache when possible."""
        mtime = os.path.getmtime(file_path)
        cached = self._column_cache.get(file_path)
        if cached and cached[0] == mtime:
            return cached[1]

        if file_path.endswith(".csv"):
            encoding = detect_encoding(file_path)
            df = pl.read_csv(
                file_path,
                has_header=True,
                n_rows=1,
                infer_schema_length=0,
                encoding=encoding,
            )
            columns = list(df.columns)
        else:
            wb = load_workbook(file_path, read_only=True)
            columns = []
            seen = set()
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                first_row = next(
                    ws.iter_rows(min_row=1, max_row=1, values_only=True), None
                )
                if first_row:
                    for cell in first_row:
                        if cell is not None:
                            col_name = str(cell).strip()
                            if col_name not in seen:
                                seen.add(col_name)
                                columns.append(col_name)
            wb.close()

        self._column_cache[file_path] = (mtime, columns)
        return columns
