import os

import polars as pl
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QHBoxLayout, QLabel, QWidget

from utils.encoding import detect_encoding


class InfoCard(QWidget):
    """Lightweight info group without border or background.

    Styles are controlled by QSS via objectName and property selectors.
    """

    def __init__(self, title):
        super().__init__()
        self.setObjectName("infoCard")
        self._value_labels = {}
        self.init_ui(title)

    def init_ui(self, title):
        # Horizontal layout: title on the left, all info rows side by side.
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        title_label = QLabel(title.upper())
        title_label.setObjectName("sectionLabel")
        layout.addWidget(title_label)

        self.content_layout = QHBoxLayout()
        self.content_layout.setSpacing(6)
        layout.addLayout(self.content_layout)

    def add_info_row(self, key, label_text):
        row = QWidget()
        row_layout = QHBoxLayout(row)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(2)

        lbl = QLabel(label_text)
        lbl.setObjectName("infoLabel")

        val = QLabel("-")
        val.setObjectName("infoValue")
        val.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        row_layout.addWidget(lbl)
        row_layout.addWidget(val)

        self.content_layout.addWidget(row)
        self._value_labels[key] = val
        return val

    def set_value(self, key, value):
        if key in self._value_labels:
            self._value_labels[key].setText(str(value))


class StatusCard(QWidget):
    """Status display area. Status color is controlled via the 'status' property."""

    def __init__(self):
        super().__init__()
        self.setObjectName("statusCard")
        self.init_ui()

    def init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        title = QLabel("STATUS")
        title.setObjectName("sectionLabel")
        layout.addWidget(title)

        self.status_label = QLabel("Ready")
        self.status_label.setWordWrap(False)
        self.status_label.setObjectName("statusValue")
        self.status_label.setProperty("status", "ready")
        layout.addWidget(self.status_label)

        self.formula_label = QLabel("")
        self.formula_label.setWordWrap(False)
        self.formula_label.setObjectName("formulaLabel")
        self.formula_label.hide()
        layout.addWidget(self.formula_label)

    def _set_status(self, text, status_type):
        self.status_label.setText(text)
        self.status_label.setProperty("status", status_type)
        self.status_label.style().unpolish(self.status_label)
        self.status_label.style().polish(self.status_label)

    def set_ready(self, formula_count=0):
        self._set_status("Ready - Click Execute to start", "ready")
        if formula_count > 0:
            self.formula_label.setText(
                f"Detected {formula_count} formula column(s) - will be auto-restored"
            )
            self.formula_label.show()
        else:
            self.formula_label.hide()

    def set_running(self):
        self._set_status("Running...", "running")

    def set_done(self):
        self._set_status("Update completed successfully", "done")

    def set_error(self, msg=""):
        display = f"Error: {msg}" if msg else "Error occurred"
        self._set_status(display, "error")


class InfoPanel(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.source_card = InfoCard("Source File Info")
        self.source_rows_label = self.source_card.add_info_row("rows", "Rows")
        self.source_cols_label = self.source_card.add_info_row(
            "cols", "Columns"
        )
        self.source_size_label = self.source_card.add_info_row("size", "Size")
        layout.addWidget(self.source_card)

        self.target_card = InfoCard("Target File Info")
        self.target_rows_label = self.target_card.add_info_row("rows", "Rows")
        self.target_cols_label = self.target_card.add_info_row(
            "cols", "Columns"
        )
        self.target_size_label = self.target_card.add_info_row("size", "Size")
        layout.addWidget(self.target_card)

        layout.addStretch(1)

        self.status = StatusCard()
        layout.addWidget(self.status)

    def _update_file_info(self, file_path, rows_label, cols_label, size_label):
        """Update file size first, then try to read row/column counts efficiently."""
        # File size is independent of parsing; always attempt to show it.
        try:
            size_mb = os.path.getsize(file_path) / (1024 * 1024)
            size_label.setText(f"{size_mb:.1f} MB")
        except OSError:
            size_label.setText("Error")

        # Row and column counts are obtained without loading all data.
        try:
            if file_path.endswith(".csv"):
                from utils.encoding import detect_encoding

                encoding = detect_encoding(file_path)

                # Get column count from header only
                header_df = pl.read_csv(
                    file_path,
                    has_header=True,
                    encoding=encoding,
                    n_rows=0,
                    infer_schema_length=0,
                )
                cols = len(header_df.columns)

                # Count rows without loading data into memory
                row_count_df = (
                    pl.scan_csv(
                        file_path,
                        has_header=True,
                        encoding=encoding,
                        infer_schema_length=0,
                    )
                    .select(pl.len())
                    .collect()
                )
                rows = row_count_df[0, 0]

                rows_label.setText(f"{rows:,}")
                cols_label.setText(str(cols))
            else:
                from openpyxl import load_workbook

                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                rows = ws.max_row
                cols = ws.max_column
                wb.close()

                rows_label.setText(f"{rows:,}")
                cols_label.setText(str(cols))
        except Exception:
            rows_label.setText("Error")
            cols_label.setText("Error")

    def update_source_info(self, file_path):
        self._update_file_info(
            file_path,
            self.source_rows_label,
            self.source_cols_label,
            self.source_size_label,
        )

    def update_target_info(self, file_path):
        self._update_file_info(
            file_path,
            self.target_rows_label,
            self.target_cols_label,
            self.target_size_label,
        )
