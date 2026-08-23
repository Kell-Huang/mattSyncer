import csv

from openpyxl import load_workbook
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QComboBox,
    QCompleter,
    QHBoxLayout,
    QLabel,
    QVBoxLayout,
    QWidget,
)

from core.sku_detector import SkuDetector
from utils.encoding import detect_encoding


class SkuPanel(QWidget):
    sku_confirmed = Signal(str, str)
    detection_warning = Signal(str)

    def __init__(self):
        super().__init__()
        self.sku_detector = SkuDetector()
        self.source_sku_col = ""
        self.target_sku_col = ""
        self._source_items = []
        self._target_items = []
        self.init_ui()

    def init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        # Source SKU column
        src_layout = QVBoxLayout()
        self.src_conf = QLabel("Waiting for files...")
        self.src_conf.setObjectName("skuStatusLabel")  # Style handled by QSS
        self.src_combo = QComboBox()
        self.src_combo.setEditable(True)
        self.src_combo.setInsertPolicy(QComboBox.NoInsert)
        self.src_combo.setFixedHeight(30)
        self.src_combo.setMaxVisibleItems(0)
        self.src_combo.activated.connect(self._on_source_activated)
        src_layout.addWidget(self.src_conf)
        src_layout.addWidget(self.src_combo)
        layout.addLayout(src_layout)

        # Target SKU column
        tgt_layout = QVBoxLayout()
        self.tgt_conf = QLabel("Waiting for files...")
        self.tgt_conf.setObjectName("skuStatusLabel")  # Style handled by QSS
        self.tgt_combo = QComboBox()
        self.tgt_combo.setEditable(True)
        self.tgt_combo.setInsertPolicy(QComboBox.NoInsert)
        self.tgt_combo.setFixedHeight(30)
        self.tgt_combo.setMaxVisibleItems(0)
        self.tgt_combo.activated.connect(self._on_target_activated)
        tgt_layout.addWidget(self.tgt_conf)
        tgt_layout.addWidget(self.tgt_combo)
        layout.addLayout(tgt_layout)

    def detect_columns(self, source_path, target_path):
        try:
            # --- Read source column names (header only) ---
            if source_path.endswith(".csv"):
                encoding = detect_encoding(source_path)
                with open(
                    source_path, "r", encoding=encoding, newline=""
                ) as f:
                    first_line = f.readline()
                    src_cols = list(next(csv.reader([first_line]), []))
            else:
                wb = load_workbook(source_path, read_only=True)
                all_src_cols = []
                seen = set()
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    first_row = next(
                        ws.iter_rows(min_row=1, max_row=1, values_only=True),
                        None,
                    )
                    if first_row:
                        for cell in first_row:
                            if cell is not None:
                                col_name = str(cell).strip()
                                if col_name not in seen:
                                    seen.add(col_name)
                                    all_src_cols.append(col_name)
                wb.close()
                src_cols = all_src_cols

            src_cols = [self._clean_column_name(c) for c in src_cols]
            self._source_items = src_cols
            src_res = self.sku_detector.detect_sku_column(src_cols)

            self._setup_combo(self.src_combo, src_cols, src_res)
            if src_res:
                self.src_combo.setCurrentText(src_res[0])
                self.source_sku_col = src_res[0]
                self.src_conf.setText("● Detected")
            else:
                self.src_conf.setText("⚠ Not detected")
                self.detection_warning.emit(
                    "Source SKU column not detected. Please select manually."
                )

            # --- Read target column names (header only) ---
            if target_path.endswith(".csv"):
                encoding = detect_encoding(target_path)
                with open(
                    target_path, "r", encoding=encoding, newline=""
                ) as f:
                    first_line = f.readline()
                    tgt_cols = list(next(csv.reader([first_line]), []))
            else:
                wb = load_workbook(target_path, read_only=True)
                all_tgt_cols = []
                seen = set()
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    first_row = next(
                        ws.iter_rows(min_row=1, max_row=1, values_only=True),
                        None,
                    )
                    if first_row:
                        for cell in first_row:
                            if cell is not None:
                                col_name = str(cell).strip()
                                if col_name not in seen:
                                    seen.add(col_name)
                                    all_tgt_cols.append(col_name)
                wb.close()
                tgt_cols = all_tgt_cols

            tgt_cols = [self._clean_column_name(c) for c in tgt_cols]
            self._target_items = tgt_cols
            tgt_res = self.sku_detector.detect_sku_column(tgt_cols)

            self._setup_combo(self.tgt_combo, tgt_cols, tgt_res)
            if tgt_res:
                self.tgt_combo.setCurrentText(tgt_res[0])
                self.target_sku_col = tgt_res[0]
                self.tgt_conf.setText("● Detected")
            else:
                self.tgt_conf.setText("⚠ Not detected")
                self.detection_warning.emit(
                    "Target SKU column not detected. Please select manually."
                )

            if self.source_sku_col and self.target_sku_col:
                self.sku_confirmed.emit(
                    self.source_sku_col, self.target_sku_col
                )
        except Exception as e:
            self.src_conf.setText("⚠ Error")
            self.detection_warning.emit(f"SKU detection error: {e}")

    def _setup_combo(self, combo, items, detected_list):
        """Configure editable combo with completer for search filtering."""
        combo.clear()
        combo.addItems(items)

        completer = QCompleter(items)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains)
        # Popup styling is handled globally by QSS
        combo.setCompleter(completer)

    def _is_valid_column(self, text, items):
        """Check if the entered text is a valid column name in the file."""
        return text in items

    def _on_source_activated(self, index):
        """Called when user selects an item from the dropdown or presses Enter."""
        text = self.src_combo.currentText().strip()
        if self._is_valid_column(text, self._source_items):
            self.source_sku_col = text
            if self.source_sku_col and self.target_sku_col:
                self.sku_confirmed.emit(
                    self.source_sku_col, self.target_sku_col
                )
        else:
            self.detection_warning.emit(
                f"Source SKU column '{text}' not found in file."
            )

    def _on_target_activated(self, index):
        """Called when user selects an item from the dropdown or presses Enter."""
        text = self.tgt_combo.currentText().strip()
        if self._is_valid_column(text, self._target_items):
            self.target_sku_col = text
            if self.source_sku_col and self.target_sku_col:
                self.sku_confirmed.emit(
                    self.source_sku_col, self.target_sku_col
                )
        else:
            self.detection_warning.emit(
                f"Target SKU column '{text}' not found in file."
            )

    def get_sku_columns(self):
        return self.src_combo.currentText(), self.tgt_combo.currentText()

    def validate(self):
        s = self.src_combo.currentText()
        t = self.tgt_combo.currentText()
        if s and t:
            if self._is_valid_column(
                s, self._source_items
            ) and self._is_valid_column(t, self._target_items):
                self.source_sku_col = s
                self.target_sku_col = t
                return True
        return False

    @staticmethod
    def _clean_column_name(col):
        """Remove leading UTF-8 BOM and surrounding whitespace from a column name."""
        return str(col).lstrip("\ufeff").strip()
